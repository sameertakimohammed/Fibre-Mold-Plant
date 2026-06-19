"""Month End Report — a faithful rebuild of the plant's emailed management template.

For years the plant has emailed an "End of Month Report - Fibre Mold Plant"
workbook: a 10-section stock & materials summary (diesel reading, goods produced,
balance stock by colour, toner, label brands on hand / used / received, pallets
local vs export, bales used / purchased). The dashboard's other reports are
production-KPI oriented; this module reproduces *that* management template so the
same document the plant has always sent can now be generated straight from the
system.

Data comes from the MonthlyStock row for the period:
  * `MonthlyStock.detail` (the full parsed 10-section payload) is rendered
    verbatim when present — every historical month imported from the old emails
    has it.
  * When `detail` is absent (a month captured only via the dashboard's stock
    form), sections fall back to the flat digest columns, and "Goods Produced"
    is computed from the daily production shifts instead.

A "Goods Produced (per daily tracker)" cross-check column is always added from
ProductionShift sums, and a deliveries summary closes the report.

    build_monthend_xlsx(db, start, end[, period_label]) -> (bytes, filename)
    build_monthend_pdf(db, start, end[, period_label])  -> (bytes, filename)
"""
import io
from datetime import date, timedelta

from sqlalchemy.orm import Session

from ..models.production import ProductionShift
from ..models.operations import Delivery, MonthlyStock
from .report_data import report_filename, collect_report_data

# Product columns, in the order the plant's template lists them.
PROD_KEYS = ["30s_small", "30s_large", "20s_normal", "12s_normal", "12s_hf",
             "12s_ff", "4cup", "2cup", "12n_labels", "12ff_labels"]
PROD_LABELS = ["30's Small", "30's Large", "20's Normal", "12's Normal",
               "12's Half Face", "12's Full Face", "4's Cup", "2's Cup",
               "12's Normal w/Labels", "12's Full Face w/Labels"]
# Map the daily-tracker production columns onto the report's product keys (the
# tracker has no separate "with labels" split, so those stay tracker-blank).
SHIFT_TO_KEY = {"p30s": "30s_small", "p30l": "30s_large", "p20n": "20s_normal",
                "p12n": "12s_normal", "p12hf": "12s_hf", "p12ff": "12s_ff",
                "p4cup": "4cup", "p2cup": "2cup"}
TONER_KEYS = ["green_liquid", "pearlscent_yellow", "toner_suit_yellow", "pigment_green_powder"]
TONER_LABELS = ["Green Liquid Colourant", "Pearlscent Pigment Yellow",
                "Toner Suit Yellow", "Pigment Green Powder"]

MONTHS = ["", "January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


def _period_of(start: date | None, end: date | None) -> str:
    d = start or end or date.today()
    return f"{d.year:04d}-{d.month:02d}"


def _month_title(period: str) -> str:
    y, m = period.split("-")
    return f"{MONTHS[int(m)]} {y}"


def collect_monthend(db: Session, start: date | None, end: date | None) -> dict:
    """Assemble everything the Month End Report needs for the period."""
    period = _period_of(start, end)
    y, m = (int(x) for x in period.split("-"))
    p_start = date(y, m, 1)
    p_end = date(y + (m // 12), (m % 12) + 1, 1)

    stock = db.query(MonthlyStock).filter(
        MonthlyStock.period == period,
        MonthlyStock.deleted_at.is_(None),
    ).first()
    detail = (stock.detail if stock else None) or {}

    # Goods produced per the daily tracker (cross-check / fallback).
    shifts = db.query(ProductionShift).filter(
        ProductionShift.deleted_at.is_(None),
        ProductionShift.work_date >= p_start,
        ProductionShift.work_date < p_end,
    ).all()
    tracker_goods = {k: 0.0 for k in PROD_KEYS}
    for s in shifts:
        for col, key in SHIFT_TO_KEY.items():
            tracker_goods[key] += getattr(s, col, 0) or 0

    goods = detail.get("goods_produced") or {}
    if not goods and stock is None:
        goods = dict(tracker_goods)  # nothing reported -> use the tracker totals

    deliveries = db.query(Delivery).filter(
        Delivery.deleted_at.is_(None),
        Delivery.work_date >= p_start,
        Delivery.work_date < p_end,
    ).order_by(Delivery.work_date).all()

    # Scalars come from the editable digest columns when a stock row exists (so a
    # manual correction via the Stock & Bales form is honoured), falling back to
    # the parsed detail payload otherwise. The richer breakdowns (goods produced,
    # balance stock by colour, toner, label brands, pallet local/export split)
    # live only in `detail`. Pallet total is overridden by the editable column.
    def scalar(col, key):
        if stock is not None:
            return getattr(stock, col, 0) or 0
        return detail.get(key, 0)

    pallets = dict(detail.get("pallets") or {})
    if stock is not None:
        pallets["total"] = stock.pallets_wrapped or pallets.get("total", 0)
    elif "total" not in pallets:
        pallets["total"] = 0

    # Headline production KPIs for the month, and the prior month for a
    # month-over-month comparison in the executive summary. Reuses the same
    # collector the production-KPI report/dashboard use, so figures agree.
    last_day = p_end - timedelta(days=1)
    metrics = collect_report_data(db, p_start, last_day, "Monthly").metrics
    prev_end = p_start - timedelta(days=1)
    prev_start = prev_end.replace(day=1)
    prev_metrics = collect_report_data(db, prev_start, prev_end, "Monthly").metrics
    prev_label = _month_title(_period_of(prev_start, prev_end))

    def _delta(cur, prev):
        """Percent change vs prior month, or None when not comparable."""
        if not prev:
            return None
        return round((cur - prev) / prev * 100, 1)

    deltas = {
        "total_qty": _delta(metrics["total_qty"], prev_metrics["total_qty"]),
        "fuel_eff": _delta(metrics["fuel_eff"], prev_metrics["fuel_eff"]),
        "downtime_pct": _delta(metrics["downtime_pct"], prev_metrics["downtime_pct"]),
        "repulp_rate": _delta(metrics["repulp_rate"], prev_metrics["repulp_rate"]),
    }

    return {
        "period": period,
        "title": _month_title(period),
        "span": (start or p_start, end or (p_end)),
        "stock": stock,
        "detail": detail,
        "goods": goods,
        "tracker_goods": tracker_goods,
        "shift_count": len(shifts),
        "deliveries": deliveries,
        "metrics": metrics,
        "prev_metrics": prev_metrics,
        "prev_label": prev_label,
        "deltas": deltas,
        "diesel_eom": scalar("diesel_eom", "diesel_eom"),
        "bales_used": scalar("bales_used", "bales_used"),
        "bales_purchased": scalar("bales_purchased", "bales_purchased"),
        "pallets": pallets,
    }


# ===========================================================================
# XLSX
# ===========================================================================
def build_monthend_xlsx(db: Session, start: date | None = None,
                        end: date | None = None, period_label: str = "MonthEnd") -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    d = collect_monthend(db, start, end)
    detail = d["detail"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Month End Report"

    AMBER = PatternFill("solid", fgColor="F5A623")
    HEAD = PatternFill("solid", fgColor="222A33")
    SUBHEAD = PatternFill("solid", fgColor="EEF1F4")
    head_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    sec_font = Font(bold=True, name="Arial", size=11, color="1A1205")
    bold = Font(bold=True, name="Arial", size=10)
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    R = lambda: Alignment(horizontal="right")

    # Title band
    ws.merge_cells("A1:F1")
    ws["A1"] = "End of Month Report — Fibre Mold Plant"
    ws["A1"].font = Font(bold=True, size=15, name="Arial", color="1A1205")
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Golden Manufacturers · Recycling Department   |   Month: {d['title']}"
    ws["A2"].font = Font(italic=True, color="667085", name="Arial", size=10)
    for c in ("A1", "A2"):
        ws[c].fill = AMBER if c == "A1" else PatternFill()
    ws["A1"].fill = AMBER
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    row = 4

    def section(title):
        nonlocal row
        ws.cell(row, 1, title).font = sec_font
        ws.cell(row, 1).fill = SUBHEAD
        for c in range(1, 7):
            ws.cell(row, c).fill = SUBHEAD
        row += 1

    def kv(label, value, nfmt="#,##0", unit=""):
        nonlocal row
        ws.cell(row, 1, label).font = bold
        cell = ws.cell(row, 2, value)
        cell.number_format = nfmt
        cell.alignment = R()
        if unit:
            ws.cell(row, 3, unit)
        row += 1

    # --- Month at a Glance (executive summary) ---
    m = d["metrics"]
    dl = d["deltas"]
    mute = Font(color="667085", name="Arial", size=9)

    def _delta_text(key):
        v = dl.get(key)
        if v is None:
            return ""
        arrow = "▲" if v > 0 else ("▼" if v < 0 else "▬")
        return f"{arrow} {abs(v):g}% vs {d['prev_label']}"

    section("Month at a Glance")
    hdr = ["Metric", "This month", "", "vs previous month"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row, i, h); c.fill = HEAD; c.font = head_font
    row += 1
    summary_rows = [
        ("Trays produced", m["total_qty"], "#,##0", "trays", "total_qty"),
        ("Active production days", m["active_days"], "0", "days", None),
        ("Average trays / day", m["avg_per_day"], "#,##0", "trays", None),
        ("Diesel burned", m["total_fuel"], "#,##0", "L", None),
        ("Fuel efficiency", m["fuel_eff"], "#,##0.0", "L / 1k trays", "fuel_eff"),
        ("Downtime", m["total_downtime_hrs"], "#,##0.0", "hrs", None),
        ("Downtime rate", m["downtime_pct"], "#,##0.0", "% of sched.", "downtime_pct"),
        ("Trays re-pulped", m["total_repulped"], "#,##0", "trays", None),
        ("Re-pulp / reject rate", m["repulp_rate"], "#,##0.0", "%", "repulp_rate"),
        ("Deliveries", m["delivery_count"], "0", "loads", None),
    ]
    for label, val, nfmt, unit, dkey in summary_rows:
        ws.cell(row, 1, label).font = bold
        cval = ws.cell(row, 2, val); cval.number_format = nfmt; cval.alignment = R()
        ws.cell(row, 3, unit).font = mute
        if dkey:
            ws.cell(row, 4, _delta_text(dkey)).font = mute
        row += 1
    row += 1

    # --- Plant Manager's Commentary (AI, optional) ---
    from . import ai
    commentary = ai.generate_commentary(d)
    if commentary:
        section("Plant Manager's Commentary")
        for line in commentary.splitlines():
            line = line.strip()
            if not line:
                continue
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row, 1, line)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(name="Arial", size=10, italic=line.lower().startswith(("recommended", "actions")))
            row += 1
        row += 1

    # 1 — Diesel
    section("1.  Diesel Reading at month end (dipstick)")
    kv("Diesel on hand", d["diesel_eom"], "#,##0", "Litres")
    row += 1

    # 2 — Goods produced (reported + tracker cross-check)
    section("2.  Goods Produced during the month")
    hdr = ["Product", "Reported (month-end)", "Per daily tracker"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row, i, h); c.fill = HEAD; c.font = head_font; c.border = border
        c.alignment = Alignment(horizontal="center")
    row += 1
    for key, label in zip(PROD_KEYS, PROD_LABELS):
        ws.cell(row, 1, label).border = border
        c1 = ws.cell(row, 2, d["goods"].get(key, 0)); c1.number_format = "#,##0"; c1.border = border; c1.alignment = R()
        tv = d["tracker_goods"].get(key, 0)
        c2 = ws.cell(row, 3, tv if tv else None); c2.number_format = "#,##0"; c2.border = border; c2.alignment = R()
        row += 1
    row += 1

    # 3 — Balance stock by colour
    section("3.  Balance Stock")
    bal = detail.get("balance_stock") or {}
    by_colour = bal.get("by_colour") or {}
    total = bal.get("total") or {}
    if by_colour or total:
        ws.cell(row, 1, "Colour").fill = HEAD; ws.cell(row, 1).font = head_font; ws.cell(row, 1).border = border
        for i, label in enumerate(PROD_LABELS, 2):
            c = ws.cell(row, i, label); c.fill = HEAD; c.font = head_font; c.border = border
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1
        for colour, vals in by_colour.items():
            ws.cell(row, 1, colour).border = border
            for i, key in enumerate(PROD_KEYS, 2):
                c = ws.cell(row, i, vals.get(key, 0) or None); c.number_format = "#,##0"; c.border = border; c.alignment = R()
            row += 1
        if total:
            ws.cell(row, 1, "Total").font = bold; ws.cell(row, 1).border = border
            for i, key in enumerate(PROD_KEYS, 2):
                c = ws.cell(row, i, total.get(key, 0) or None); c.number_format = "#,##0"; c.font = bold; c.border = border; c.alignment = R()
            row += 1
    else:
        kv("30's balance", getattr(d["stock"], "bal_30s", 0) if d["stock"] else 0)
        kv("12's Normal balance", getattr(d["stock"], "bal_12n", 0) if d["stock"] else 0)
        kv("12's Full Face balance", getattr(d["stock"], "bal_12ff", 0) if d["stock"] else 0)
    row += 1

    # 4 — Toner / colourants
    section("4.  Toner & Colourants (Kg)")
    toner = detail.get("toner") or {}
    for key, label in zip(TONER_KEYS, TONER_LABELS):
        kv(label, toner.get(key, 0), "#,##0.##", "Kg")
    row += 1

    # 5 / 6 / 7 — Label brands
    def label_section(title, items):
        nonlocal row
        section(title)
        if not items:
            ws.cell(row, 1, "—"); row += 1; return
        for it in items:
            ws.cell(row, 1, it.get("brand", "")).font = bold
            c = ws.cell(row, 2, it.get("qty", 0)); c.number_format = "#,##0"; c.alignment = R()
            ws.cell(row, 3, "pcs")
            row += 1
    label_section("5.  Label Stickers on hand", detail.get("labels_on_hand", []))
    row += 1
    label_section("6.  Label Stickers used", detail.get("labels_used", []))
    row += 1
    label_section("7.  Label / reel stock received", detail.get("stock_received", []))
    row += 1

    # 8 — Pallets
    section("8.  Pallets wrapped during the month")
    p = d["pallets"]
    kv("Total pallets wrapped", p.get("total", 0), "#,##0", "Pallet")
    kv("Local (single wrap)", p.get("local", 0), "#,##0")
    kv("Export (double wrap)", p.get("export", 0), "#,##0")
    row += 1

    # 9 / 10 — Bales
    section("9.  Bales")
    kv("Bales used (per consumption book)", d["bales_used"], "#,##0.##")
    kv("Bales purchased", d["bales_purchased"], "#,##0.##")
    row += 2

    # Deliveries summary
    section(f"Deliveries during the month ({len(d['deliveries'])})")
    dh = ["Date", "Customer", "30's", "12's Normal", "12's Full Face", "Pallets"]
    for i, h in enumerate(dh, 1):
        c = ws.cell(row, i, h); c.fill = HEAD; c.font = head_font; c.border = border
    row += 1
    for dv in d["deliveries"]:
        vals = [dv.work_date.isoformat(), dv.company, dv.tray30, dv.tray12n, dv.tray12ff, dv.pallets]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v); c.border = border
            if isinstance(v, (int, float)):
                c.number_format = "#,##0"; c.alignment = R()
        row += 1

    widths = [30, 16, 14, 14, 14, 12, 12, 12, 12, 12, 12]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), report_filename("xlsx", start, end, "MonthEnd")


# ===========================================================================
# PDF
# ===========================================================================
def build_monthend_pdf(db: Session, start: date | None = None,
                       end: date | None = None, period_label: str = "MonthEnd") -> tuple[bytes, str]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, KeepTogether)
    from xml.sax.saxutils import escape

    AMBER = colors.HexColor("#F5A623")
    INK = colors.HexColor("#1A1205")
    DARK = colors.HexColor("#222A33")
    MUTE = colors.HexColor("#667085")
    LINE = colors.HexColor("#D0D5DD")
    ZEBRA = colors.HexColor("#F7F8FA")

    d = collect_monthend(db, start, end)
    detail = d["detail"]
    ss = getSampleStyleSheet()
    st = {
        "title": ParagraphStyle("t", parent=ss["Title"], fontName="Helvetica-Bold",
                                fontSize=18, textColor=colors.white, leading=22),
        "sub": ParagraphStyle("s", parent=ss["Normal"], fontSize=9.5, textColor=colors.white, leading=13),
        "h2": ParagraphStyle("h2", parent=ss["Heading2"], fontName="Helvetica-Bold",
                             fontSize=11.5, textColor=INK, spaceBefore=10, spaceAfter=4),
        "cell": ParagraphStyle("c", parent=ss["Normal"], fontSize=8, leading=10),
        "foot": ParagraphStyle("f", parent=ss["Normal"], fontSize=7.5, textColor=MUTE),
    }

    def num(v):
        if not v:
            return ""
        return f"{v:,.0f}" if float(v).is_integer() else f"{v:,.2f}"

    def header():
        inner = [[Paragraph("End of Month Report — Fibre Mold Plant", st["title"])],
                 [Paragraph(f"Golden Manufacturers · Recycling Department &nbsp;|&nbsp; Month: {d['title']}", st["sub"])]]
        t = Table(inner, colWidths=[257 * mm])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), AMBER),
                               ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                               ("TOPPADDING", (0, 0), (0, 0), 10), ("BOTTOMPADDING", (0, -1), (-1, -1), 10)]))
        return t

    def grid(headers, rows, widths, bold_last=False):
        body = [[Paragraph(f'<font color="white"><b>{escape(str(h))}</b></font>', st["cell"]) for h in headers]]
        for r in rows:
            body.append([Paragraph(escape(str(c)), st["cell"]) for c in r])
        t = Table(body, colWidths=widths, repeatRows=1)
        style = [("BACKGROUND", (0, 0), (-1, 0), DARK),
                 ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ZEBRA]),
                 ("GRID", (0, 0), (-1, -1), 0.4, LINE),
                 ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                 ("ALIGN", (1, 1), (-1, -1), "RIGHT")]
        if bold_last:
            style.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
        t.setStyle(TableStyle(style))
        return t

    flow = [header(), Spacer(1, 8)]

    # --- Month at a Glance (executive summary) ---
    m = d["metrics"]
    dl = d["deltas"]

    def delta_cell(key):
        v = dl.get(key)
        if v is None:
            return ""
        arrow = "▲" if v > 0 else ("▼" if v < 0 else "▬")
        return f"{arrow} {abs(v):g}% vs {d['prev_label']}"

    flow.append(Paragraph("Month at a Glance", st["h2"]))
    srows = [
        ["Trays produced", num(m["total_qty"]), delta_cell("total_qty")],
        ["Active production days", num(m["active_days"]), ""],
        ["Average trays / day", num(m["avg_per_day"]), ""],
        ["Diesel burned (L)", num(m["total_fuel"]), ""],
        ["Fuel efficiency (L / 1k trays)", num(m["fuel_eff"]), delta_cell("fuel_eff")],
        ["Downtime (hrs)", num(m["total_downtime_hrs"]), ""],
        ["Downtime rate (% of scheduled)", num(m["downtime_pct"]), delta_cell("downtime_pct")],
        ["Trays re-pulped", num(m["total_repulped"]), ""],
        ["Re-pulp / reject rate (%)", num(m["repulp_rate"]), delta_cell("repulp_rate")],
        ["Deliveries (loads)", num(m["delivery_count"]), ""],
    ]
    g = grid(["Metric", "This month", "vs previous month"], srows,
             [95 * mm, 45 * mm, 75 * mm])
    # The "vs previous" column reads as a note, not a number — left-align it.
    g.setStyle(TableStyle([("ALIGN", (2, 1), (2, -1), "LEFT")]))
    flow.append(g)
    flow.append(Spacer(1, 4))

    # --- Plant Manager's Commentary (AI, optional) ---
    from . import ai
    commentary = ai.generate_commentary(d)
    if commentary:
        body = ParagraphStyle("cmt", parent=st["cell"], fontSize=9, leading=12,
                              spaceAfter=3)
        flow.append(Paragraph("Plant Manager's Commentary", st["h2"]))
        for line in commentary.splitlines():
            line = line.strip()
            if line:
                flow.append(Paragraph(escape(line), body))
        flow.append(Spacer(1, 6))

    # 1 — diesel + 9/10 bales in a quick KPI strip
    flow.append(Paragraph("Stock summary", st["h2"]))
    flow.append(grid(["Metric", "Value"], [
        ["Diesel on hand (Litres)", num(d["diesel_eom"])],
        ["Bales used (consumption book)", num(d["bales_used"])],
        ["Bales purchased", num(d["bales_purchased"])],
        ["Pallets wrapped", num(d["pallets"].get("total", 0))],
        ["  · Local (single wrap)", num(d["pallets"].get("local", 0))],
        ["  · Export (double wrap)", num(d["pallets"].get("export", 0))],
    ], [120 * mm, 40 * mm]))

    # 2 — goods produced
    flow.append(Paragraph("Goods Produced during the month", st["h2"]))
    grows = [[lbl, num(d["goods"].get(k, 0)), num(d["tracker_goods"].get(k, 0))]
             for k, lbl in zip(PROD_KEYS, PROD_LABELS)]
    flow.append(grid(["Product", "Reported", "Per daily tracker"], grows,
                     [90 * mm, 45 * mm, 45 * mm]))

    # 3 — balance stock
    bal = detail.get("balance_stock") or {}
    by_colour, total = bal.get("by_colour") or {}, bal.get("total") or {}
    if by_colour or total:
        flow.append(Paragraph("Balance Stock", st["h2"]))
        short = ["30s S", "30s L", "20s N", "12 N", "12 HF", "12 FF", "4cup", "2cup", "12N-L", "12FF-L"]
        brows = [[c] + [num(v.get(k, 0)) for k in PROD_KEYS] for c, v in by_colour.items()]
        if total:
            brows.append(["Total"] + [num(total.get(k, 0)) for k in PROD_KEYS])
        w = [30 * mm] + [22.7 * mm] * 10
        flow.append(grid(["Colour"] + short, brows, w, bold_last=bool(total)))

    # 5/6/7 — labels
    def label_rows(items):
        return [[i.get("brand", ""), num(i.get("qty", 0))] for i in items] or [["—", ""]]
    for title, key in [("Label Stickers on hand", "labels_on_hand"),
                       ("Label Stickers used", "labels_used"),
                       ("Label / reel stock received", "stock_received")]:
        flow.append(Paragraph(title, st["h2"]))
        flow.append(grid(["Brand", "Qty"], label_rows(detail.get(key, [])), [120 * mm, 40 * mm]))

    # deliveries
    flow.append(Paragraph(f"Deliveries during the month ({len(d['deliveries'])})", st["h2"]))
    drows = [[dv.work_date.isoformat(), (dv.company or "")[:34], num(dv.tray30),
              num(dv.tray12n), num(dv.tray12ff), num(dv.pallets)] for dv in d["deliveries"]]
    flow.append(grid(["Date", "Customer", "30's", "12's Normal", "12's FF", "Pallets"],
                     drows or [["—", "", "", "", "", ""]],
                     [26 * mm, 70 * mm, 24 * mm, 28 * mm, 24 * mm, 24 * mm]))

    flow.append(Spacer(1, 10))
    flow.append(Paragraph("Generated by the Fibre Mold Plant dashboard · reproduces the plant's monthly End-of-Month Report.", st["foot"]))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=18 * mm, rightMargin=18 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
                            title=f"Fibre Mold Plant Month End Report {d['title']}", author="Golden Manufacturers")
    doc.build(flow)
    buf.seek(0)
    return buf.getvalue(), report_filename("pdf", start, end, "MonthEnd")
