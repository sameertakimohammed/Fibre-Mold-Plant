"""Excel (.xlsx) report builder.

The openpyxl workbook construction. Figures come from collect_report_data() in
report_data.py, the single source shared with the PDF/PPTX builders and the
scheduled report-email job — so every format and the emailed copy agree.

build_report_xlsx(db, start, end[, period_label]) -> (xlsx_bytes, filename)

The (db, start, end) call shape is preserved for the scheduler, which calls it
positionally.
"""
import io
from datetime import date

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .report_data import collect_report_data, report_filename
from .report_trends import collect_trend_data, deck_narrative

AMBER = "F5A623"
INK = "1A1205"
HEAD_FILL = PatternFill("solid", fgColor="222A33")
HEAD_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TITLE_FONT = Font(bold=True, size=15, name="Arial")
LABEL_FONT = Font(bold=True, name="Arial", size=10)
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 40)


def _header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def build_report_xlsx(
    db: Session,
    start: date | None = None,
    end: date | None = None,
    period_label: str = "Production",
) -> tuple[bytes, str]:
    """Build the production report workbook for the given date range.

    Returns (xlsx_bytes, filename). Soft-deleted rows are excluded so the report
    matches the dashboard KPIs.
    """
    data = collect_report_data(db, start, end, period_label)
    m = data.metrics
    shifts, deliveries = data.shifts, data.deliveries

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Fibre Mold Plant — {data.period_label} Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Golden Manufacturers · Period: {data.span}"
    ws["A2"].font = Font(italic=True, color="667085", name="Arial", size=10)

    rows = [
        ("Total trays produced", m["total_qty"], "#,##0"),
        ("Active production days", m["active_days"], "0"),
        ("Average trays / day", m["avg_per_day"], "#,##0"),
        ("Total diesel burned (L)", m["total_fuel"], "#,##0"),
        ("Fuel efficiency (L / 1k trays)", m["fuel_eff"], "#,##0.0"),
        ("Total downtime (hrs)", m["total_downtime_hrs"], "#,##0.0"),
        ("Downtime rate (% of scheduled)", m["downtime_pct"], "#,##0.0"),
        ("Trays re-pulped", m["total_repulped"], "#,##0"),
        ("Re-pulp rate (%)", m["repulp_rate"], "#,##0.0"),
        ("30's delivered", m["tray30"], "#,##0"),
        ("12's delivered", m["tray12"], "#,##0"),
        ("Pallets shipped", m["pallets"], "#,##0"),
    ]
    r = 4
    for label, val, nfmt in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = nfmt
        c.alignment = Alignment(horizontal="right")
        r += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16

    # ---- Shift detail sheet ----
    sd = wb.create_sheet("Shift Detail")
    headers = ["Date", "Shift", "Trays", "Speed", "Prod Hrs", "Sched Hrs",
               "Fuel Used (L)", "Downtime (min)", "Re-pulped", "Comment"]
    _header_row(sd, 1, headers)
    for i, s in enumerate(shifts, 2):
        vals = [s.work_date.isoformat(), s.shift.value, s.qty, s.speed, s.prod_hours,
                s.sched_hours, s.fuel_use, s.downtime_min, s.repulped, s.comment or ""]
        for j, v in enumerate(vals, 1):
            c = sd.cell(row=i, column=j, value=v)
            if isinstance(v, (int, float)):
                c.number_format = "#,##0.#"
            c.border = BORDER
    sd.freeze_panes = "A2"
    _autofit(sd)

    # ---- Deliveries sheet ----
    dl = wb.create_sheet("Deliveries")
    dheaders = ["Date", "Customer", "30's", "12's Normal", "12's Full Face", "Pallets", "Comment"]
    _header_row(dl, 1, dheaders)
    for i, d in enumerate(deliveries, 2):
        vals = [d.work_date.isoformat(), d.company, d.tray30, d.tray12n, d.tray12ff, d.pallets, d.comment or ""]
        for j, v in enumerate(vals, 1):
            c = dl.cell(row=i, column=j, value=v)
            if isinstance(v, (int, float)):
                c.number_format = "#,##0.#"
            c.border = BORDER
    dl.freeze_panes = "A2"
    _autofit(dl)

    # ---- AI Improvement Plan sheet (only when AI is enabled & returns one) ----
    narrative = deck_narrative(collect_trend_data(db, start, end))
    plan = narrative.get("improvement_plan") if narrative else None
    if plan:
        ip = wb.create_sheet("Improvement Plan")
        ip["A1"] = "AI Improvement Plan"
        ip["A1"].font = TITLE_FONT
        overall = narrative.get("overall_summary")
        if overall:
            ip["A2"] = overall
            ip["A2"].font = Font(italic=True, color="667085", name="Arial", size=10)
            ip["A2"].alignment = Alignment(wrap_text=True, vertical="top")
            ip.merge_cells("A2:D2")
            ip.row_dimensions[2].height = 30
        _header_row(ip, 4, ["#", "Issue", "Detail", "Recommended action"])
        for i, item in enumerate(plan, 1):
            vals = [i, str(item.get("title", "")), str(item.get("detail", "")),
                    str(item.get("action", ""))]
            for j, v in enumerate(vals, 1):
                c = ip.cell(row=4 + i, column=j, value=v)
                c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top",
                                        horizontal="center" if j == 1 else "left")
        ip.column_dimensions["A"].width = 5
        ip.column_dimensions["B"].width = 30
        ip.column_dimensions["C"].width = 55
        ip.column_dimensions["D"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), report_filename("xlsx", start, end, period_label)
