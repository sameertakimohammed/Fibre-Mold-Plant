"""
Import an existing monthly tracker Excel file into the database.

Usage (from the backend/ folder, with the app running or DB reachable):
    python -m app.services.import_excel /path/to/FIBER_MOLD_TRACKER_JUNE_2026.xlsx

This reads the same sheet layout as the original May tracker. Existing
date+shift rows are skipped so you can safely re-run it.

Requires: pip install openpyxl
"""
import sys
import datetime
from sqlalchemy.orm import Session
from ..core.database import SessionLocal
from ..models.production import ProductionShift, Shift


def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("m", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _shift_enum(s: str) -> Shift:
    l = (s or "").lower()
    if l.startswith("day"):
        return Shift.day
    if l.startswith("afternoon"):
        return Shift.afternoon
    return Shift.night


def import_tracker(path: str, db: Session) -> int:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    existing = {(r.work_date, r.shift) for r in db.query(ProductionShift).all()}
    added = 0
    cur_date = None
    for r in range(5, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if isinstance(a, datetime.datetime):
            cur_date = a.date()
        if not b or not str(b).lower().startswith(("day", "afternoon", "night")):
            continue
        if cur_date is None:
            continue
        sh = _shift_enum(str(b))
        if (cur_date, sh) in existing:
            continue
        existing.add((cur_date, sh))
        ps = ProductionShift(
            work_date=cur_date, shift=sh,
            qty=_num(ws.cell(r, 3).value),
            p30s=_num(ws.cell(r, 4).value), p30l=_num(ws.cell(r, 5).value),
            p20n=_num(ws.cell(r, 6).value), p12n=_num(ws.cell(r, 7).value),
            p12hf=_num(ws.cell(r, 8).value), p12ff=_num(ws.cell(r, 9).value),
            p4cup=_num(ws.cell(r, 10).value), p2cup=_num(ws.cell(r, 11).value),
            hp1=_num(ws.cell(r, 12).value), hp2=_num(ws.cell(r, 13).value),
            hp3=_num(ws.cell(r, 14).value), hp4=_num(ws.cell(r, 15).value),
            hp5=_num(ws.cell(r, 16).value), hp6=_num(ws.cell(r, 17).value),
            labelling=_num(ws.cell(r, 18).value), water_meter=_num(ws.cell(r, 19).value),
            carton_bales=_num(ws.cell(r, 20).value), speed=_num(ws.cell(r, 21).value),
            fuel_open=_num(ws.cell(r, 22).value), fuel_close=_num(ws.cell(r, 23).value),
            fuel_use=_num(ws.cell(r, 24).value),
            prod_hours=_num(ws.cell(r, 25).value), downtime_min=_num(ws.cell(r, 26).value),
            sched_hours=_num(ws.cell(r, 28).value) or 8,
            clean_min=_num(ws.cell(r, 30).value), mold_min=_num(ws.cell(r, 31).value),
            other_min=_num(ws.cell(r, 32).value), repulped=_num(ws.cell(r, 33).value),
            comment=str(ws.cell(r, 34).value or ""),
        )
        db.add(ps)
        added += 1
    db.commit()
    return added


def main():
    if len(sys.argv) < 2:
        print("Usage: python -m app.services.import_excel <tracker.xlsx>")
        sys.exit(1)
    db = SessionLocal()
    try:
        n = import_tracker(sys.argv[1], db)
        print(f"Imported {n} new shifts from {sys.argv[1]}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
