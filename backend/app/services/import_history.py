"""Historical multi-month, multi-file importer for the Fibre Mold Plant.

The plant has emailed a monthly bundle for years: a daily production tracker
plus deliveries, fuel-dip, bale and an "End of Month Report" stock record. The
original `import_excel.py` handled a single tracker only. This module adds
format-tolerant parsers for *all five* file types and a driver that walks a
directory of those files, parses every one, and merges the rows with natural-key
de-duplication so the same month appearing in several emails (or an "UPDATED"
tracker) collapses to one clean dataset.

Two entry points:

    build_history_json(src_dir, out_path)   # parse raw .xlsx -> data/history.json
    import_history(db[, json_path])          # load history.json into the DB (idempotent)

`build_history_json` is the offline step run once against the extracted email
attachments; the resulting `data/history.json` is committed and replayed on
every fresh install by `import_history`, exactly mirroring how the original May
data shipped as `tracker.json`/`support.json`.

Parsers locate structure by content (header text, section markers) rather than
fixed cell coordinates, because the layout drifted across years — the 2024
tracker header sits on row 5, the 2026 one on row 3; the month-end "Total" row
moved between revisions. Dates are carried forward across a day's shift rows and
parsed defensively (real datetimes *and* malformed strings like "01/04/026").
"""
from __future__ import annotations

import datetime
import glob
import json
import logging
import os
import re
from dataclasses import dataclass, field

logger = logging.getLogger("app.import_history")

SHIFT_DAY = "Day"
SHIFT_AFT = "Afternoon"
SHIFT_NIGHT = "Night"


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def _num(v) -> float:
    """Coerce a cell to a float, tolerating commas, stray units and #DIV/0!."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower().replace(",", "")
    s = re.sub(r"(litres?|liters?|pcs|kg|m\^?3|m|l|reels?|pallets?)\b", "", s)
    s = s.replace("m", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else 0.0


def _to_int(v) -> int:
    return int(round(_num(v)))


def _shift_enum(s) -> str:
    l = (str(s) or "").strip().lower()
    if l.startswith("day") or "7.30am" in l or "7:30am" in l:
        return SHIFT_DAY
    if l.startswith("afternoon") or "3.30pm" in l or "3:30pm" in l:
        return SHIFT_AFT
    if l.startswith("night") or "11.30pm" in l or "11:30pm" in l:
        return SHIFT_NIGHT
    return ""


def _parse_date(v, ctx_year=None, ctx_month=None):
    """Return a datetime.date or None. Accepts real datetimes and dd/mm/yy[yy]
    strings (including the malformed '01/04/026' year)."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    m = re.search(r"(\d{1,2})\s*[/.\-]\s*(\d{1,2})\s*[/.\-]\s*(\d{2,4})", s)
    if not m:
        return None
    day, mon, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if yr < 100:
        yr = 2000 + yr if yr < 70 else 1900 + yr
    elif yr < 1000:           # the '026' case -> 26 -> 2026
        yr = 2000 + (yr % 100)
    try:
        return datetime.date(yr, mon, day)
    except ValueError:
        return None


def _find_header(ws, want, max_scan=20):
    """Return the 1-based row index whose cells contain all `want` substrings
    (case-insensitive), or None."""
    want = [w.lower() for w in want]
    for r in range(1, min(ws.max_row, max_scan) + 1):
        rowtext = " ".join(
            str(ws.cell(r, c).value).lower()
            for c in range(1, min(ws.max_column, 40) + 1)
            if ws.cell(r, c).value is not None
        )
        if all(w in rowtext for w in want):
            return r
    return None


def _period_of(d: datetime.date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


# ---------------------------------------------------------------------------
# Parsers — each returns a list[dict] (or dict) of plain JSON-able values
# ---------------------------------------------------------------------------
def parse_tracker(path: str) -> list[dict]:
    """Daily production tracker -> one dict per shift row with real data."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hdr = _find_header(ws, ["date", "shift", "quantity"]) or 3
    data_start = hdr + 2
    out, cur_date = [], None
    for r in range(data_start, ws.max_row + 1):
        d = _parse_date(ws.cell(r, 1).value)
        if d:
            cur_date = d
        shift = _shift_enum(ws.cell(r, 2).value)
        if not shift or cur_date is None:
            continue
        row = {
            "date": cur_date.isoformat(), "shift": shift,
            "qty": _num(ws.cell(r, 3).value),
            "p30s": _num(ws.cell(r, 4).value), "p30l": _num(ws.cell(r, 5).value),
            "p20n": _num(ws.cell(r, 6).value), "p12n": _num(ws.cell(r, 7).value),
            "p12hf": _num(ws.cell(r, 8).value), "p12ff": _num(ws.cell(r, 9).value),
            "p4cup": _num(ws.cell(r, 10).value), "p2cup": _num(ws.cell(r, 11).value),
            "hp1": _num(ws.cell(r, 12).value), "hp2": _num(ws.cell(r, 13).value),
            "hp3": _num(ws.cell(r, 14).value), "hp4": _num(ws.cell(r, 15).value),
            "hp5": _num(ws.cell(r, 16).value), "hp6": _num(ws.cell(r, 17).value),
            "label": _num(ws.cell(r, 18).value), "water": _num(ws.cell(r, 19).value),
            "bales": _num(ws.cell(r, 20).value), "speed": _num(ws.cell(r, 21).value),
            "fuel_open": _num(ws.cell(r, 22).value), "fuel_close": _num(ws.cell(r, 23).value),
            "fuel_use": _num(ws.cell(r, 24).value),
            "prod_hrs": _num(ws.cell(r, 25).value), "down_min": _num(ws.cell(r, 26).value),
            "tot_hrs": _num(ws.cell(r, 28).value) or 8,
            "clean": _num(ws.cell(r, 30).value), "mold": _num(ws.cell(r, 31).value),
            "other": _num(ws.cell(r, 32).value), "repulp": _num(ws.cell(r, 33).value),
            "comment": str(ws.cell(r, 34).value or "").strip(),
        }
        # Drop the blank template rows that pad these sheets to 1000+ rows.
        if (row["qty"] == 0 and row["prod_hrs"] == 0 and row["down_min"] == 0
                and row["fuel_use"] == 0 and not row["comment"]):
            continue
        out.append(row)
    return out


def parse_deliveries(path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hdr = _find_header(ws, ["date", "company"]) or 1
    out, cur_date = [], None
    for r in range(hdr + 1, ws.max_row + 1):
        d = _parse_date(ws.cell(r, 1).value)
        if d:
            cur_date = d
        company = str(ws.cell(r, 2).value or "").strip()
        t30 = _num(ws.cell(r, 3).value)
        t12n = _num(ws.cell(r, 4).value)
        t12ff = _num(ws.cell(r, 5).value)
        pallets = _num(ws.cell(r, 6).value)
        comment = str(ws.cell(r, 7).value or "").strip()
        if cur_date is None:
            continue
        # Skip the month-total / summary footer rows (no real customer name, or a
        # literal "total"), and fully blank rows.
        if "total" in company.lower() or not re.search(r"[A-Za-z]", company):
            continue
        if t30 == 0 and t12n == 0 and t12ff == 0 and pallets == 0:
            continue
        out.append({
            "date": cur_date.isoformat(), "company": company or "—",
            "t30": t30, "t12n": t12n, "t12ff": t12ff, "pallets": pallets,
            "comment": comment,
        })
    return out


# Fuel-dip sheets lay three shift blocks side by side. Each tuple is
# (date_col, open_col, close_col, usage_col, note_col, shift_label), 1-based.
_FUEL_BLOCKS = [
    (1, 2, 3, 5, 9, SHIFT_DAY),       # A,B,C,E + note I
    (12, 13, 14, 16, 17, SHIFT_AFT),  # L,M,N,P + note Q
    (20, 21, 22, 24, 25, SHIFT_NIGHT),  # T,U,V,X + note Y
]


def parse_fuel(path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # Data starts at the first row whose first column is a real date.
    start = None
    for r in range(1, min(ws.max_row, 15) + 1):
        if _parse_date(ws.cell(r, 1).value) and isinstance(
                ws.cell(r, 1).value, (datetime.datetime, datetime.date)):
            start = r
            break
    if start is None:
        return []
    out = []
    for r in range(start, ws.max_row + 1):
        for dc, oc, cc, uc, nc, shift in _FUEL_BLOCKS:
            d = _parse_date(ws.cell(r, dc).value)
            if not d:
                continue
            o = _num(ws.cell(r, oc).value)
            c = _num(ws.cell(r, cc).value)
            # Prefer the sheet's explicit "actual usage"; fall back to the dip
            # delta but never report negative (a tank refill makes close > open).
            usage = _num(ws.cell(r, uc).value)
            if usage == 0 and o and c:
                usage = max(0.0, o - c)
            note = str(ws.cell(r, nc).value or "").strip()
            received = 0.0
            mrec = re.search(r"received\s*([\d,]+)\s*l", note, re.I)
            if mrec:
                received = _num(mrec.group(1))
            if o == 0 and c == 0 and usage == 0 and received == 0:
                continue
            out.append({
                "date": d.isoformat(), "shift": shift,
                "open": o, "close": c, "usage": usage,
                "received": received, "note": note,
            })
    return out


def parse_bales(path: str) -> list[dict]:
    """Bale daily sheet -> goods-received rows. Only rows carrying a GRN# are
    treated as receipts (the rest is per-shift usage already in the tracker)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hdr = _find_header(ws, ["goods received", "weight"]) or _find_header(ws, ["bale", "weight"]) or 4
    out, cur_date = [], None
    for r in range(hdr + 1, ws.max_row + 1):
        d = _parse_date(ws.cell(r, 1).value)
        if d:
            cur_date = d
        grn = str(ws.cell(r, 2).value or "").strip()
        weight = _num(ws.cell(r, 3).value)
        qty = _num(ws.cell(r, 4).value)
        # Only genuine goods-received rows: a GRN reference *and* a weight/qty.
        # (The footer of these sheets holds free-text notes in the GRN column.)
        if not grn or cur_date is None or (weight == 0 and qty == 0):
            continue
        out.append({
            "date": cur_date.isoformat(), "grn": grn,
            "weight_kg": weight, "quantity": qty,
        })
    return out


def parse_month_end(path: str, period: str) -> dict:
    """Parse the 10-section 'End of Month Report' into a structured detail dict
    plus the flat digest fields used by the MonthlyStock row."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Locate each numbered section by the integer marker in column A. Column A in
    # these sheets only ever holds the 1..10 section numbers (and the odd label),
    # so the first occurrence of each is the marker. We require *some* content in
    # B or C on that row so a stray value can't be mistaken for a marker — but we
    # don't require B specifically, because section 2's label cell is sometimes
    # blank (the product headers sit in C onward).
    sec: dict[int, int] = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, int) and not isinstance(v, bool) and 1 <= v <= 10 and v not in sec:
            if ws.cell(r, 2).value is not None or ws.cell(r, 3).value is not None:
                sec[v] = r

    def row_vals(r, cols):
        return [ws.cell(r, c).value for c in cols]

    PROD_COLS = list(range(3, 13))  # C..L : 30s small..12ff w/labels
    PROD_KEYS = ["30s_small", "30s_large", "20s_normal", "12s_normal",
                 "12s_hf", "12s_ff", "4cup", "2cup", "12n_labels", "12ff_labels"]

    detail: dict = {"period": period}

    # 1 - diesel reading
    if 1 in sec:
        detail["diesel_eom"] = _num(ws.cell(sec[1], 3).value)
    # 2 - goods produced (value row is 3 rows below the marker: marker, dims, "Approx.", values)
    if 2 in sec:
        for off in (3, 2, 4):
            vals = row_vals(sec[2] + off, PROD_COLS)
            if any(isinstance(x, (int, float)) for x in vals):
                detail["goods_produced"] = {
                    k: _num(v) for k, v in zip(PROD_KEYS, vals)}
                break
    # 3 - balance stock matrix (rows labelled by colour in column B until 'Total')
    if 3 in sec:
        colours = {}
        total = {}
        r = sec[3] + 2
        end = sec.get(4, ws.max_row)
        while r < end:
            label = str(ws.cell(r, 2).value or "").strip()
            if label:
                vals = {k: _num(v) for k, v in zip(PROD_KEYS, row_vals(r, PROD_COLS))}
                if label.lower().startswith("total"):
                    total = vals
                else:
                    colours[label] = vals
            r += 1
        detail["balance_stock"] = {"by_colour": colours, "total": total}
    # 4 - toner (KG row beneath the marker)
    if 4 in sec:
        tcols = list(range(3, 7))
        tkeys = ["green_liquid", "pearlscent_yellow", "toner_suit_yellow", "pigment_green_powder"]
        detail["toner"] = {k: _num(v) for k, v in zip(tkeys, row_vals(sec[4] + 1, tcols))}
    # 5 / 6 / 7 - label brands (qty in C, brand in E)
    def label_block(marker):
        items, end = [], None
        if marker not in sec:
            return items
        nxt = min((sec[n] for n in sec if sec[n] > sec[marker]), default=ws.max_row + 1)
        for r in range(sec[marker], nxt):
            qty = ws.cell(r, 3).value
            brand = str(ws.cell(r, 5).value or "").strip()
            if brand:
                items.append({"brand": brand, "qty": _num(qty), "raw": str(qty) if qty is not None else ""})
        return items
    detail["labels_on_hand"] = label_block(5)
    detail["labels_used"] = label_block(6)
    detail["stock_received"] = label_block(7)
    # 8 - pallets wrapped
    if 8 in sec:
        r = sec[8]
        detail["pallets"] = {
            "total": _num(ws.cell(r, 3).value),
            "local": _num(ws.cell(r, 7).value),
            "export": _num(ws.cell(r, 10).value),
            "single_wrap": _num(ws.cell(r + 1, 9).value),
            "double_wrap": _num(ws.cell(r + 1, 12).value),
        }
    # 9 / 10 - bales used / purchased
    if 9 in sec:
        detail["bales_used"] = _num(ws.cell(sec[9], 3).value)
    if 10 in sec:
        detail["bales_purchased"] = _num(ws.cell(sec[10], 3).value)

    bal = detail.get("balance_stock", {}).get("total", {})
    digest = {
        "period": period,
        "diesel_eom": detail.get("diesel_eom", 0),
        "bal_30s": bal.get("30s_small", 0),
        "bal_12n": bal.get("12s_normal", 0),
        "bal_12ff": bal.get("12s_ff", 0),
        "bal_12nl": bal.get("12n_labels", 0),
        "pallets_wrapped": detail.get("pallets", {}).get("total", 0),
        "bales_used": detail.get("bales_used", 0),
        "bales_purchased": detail.get("bales_purchased", 0),
        "labels_used": sum(i["qty"] for i in detail.get("labels_used", [])),
        "detail": detail,
    }
    return digest


# ---------------------------------------------------------------------------
# Driver — walk a directory of month files into one consolidated dataset
# ---------------------------------------------------------------------------
# The "End of Month Report" stock files are a single-period digest and several
# are generically named ("Month End Report_Dec.xlsx") with no year inside, so we
# map each source folder/file to its period explicitly. Trackers, deliveries,
# fuel and bale files instead get their period from the dates inside them.
MONTH_END_PERIODS = {
    "month end report_aug": "2024-08",
    "month end report_sep": "2024-09",
    "month end report_oct": "2024-10",
    "month end report_nov": "2024-11",
    "month end report_dec": "2024-12",
    "month end report_feb 2025": "2025-02",
    "month end record(aug 2025)": "2025-08",
    "month end record(sep 2025)": "2025-09",
    "month end record(october 2025)": "2025-10",
    "month end record(november 2025)": "2025-11",
    "month end stock record(january 2026)": "2026-01",
    "month end stock record(febuary 2026)": "2026-02",
    "month end stock record(march 2026)": "2026-03",
    "month end stock record(april 2026)": "2026-04",
    "month end stock record(may 2026)": "2026-05",
}

MIN_PERIOD = "2024-08"  # requested import window starts here


def _classify(name: str) -> str:
    n = name.lower()
    if "deliver" in n:
        return "deliveries"
    if "fuel" in n:
        return "fuel"
    if "month end" in n and ("record" in n or "report" in n or "stock" in n):
        return "month_end"
    if "bale" in n or "bail" in n:
        return "bales"
    if "tracker" in n:
        return "tracker"
    # Some months ship the tracker under a "...REPORT" name (e.g. "FIBRE MOLD OCT
    # REPORT^.xlsx"); the daily-usage tracker is the only file that carries the
    # "fibre/fiber mold" brand without one of the keywords above.
    if "fibre mold" in n or "fiber mold" in n:
        return "tracker"
    return ""


def _month_end_period(path: str) -> str | None:
    stem = os.path.splitext(os.path.basename(path))[0].lower()
    stem = re.sub(r"\s*\(\d+\)$", "", stem).strip(" .")  # drop " (2)" copy suffix
    for key, period in MONTH_END_PERIODS.items():
        if stem.startswith(key):
            return period
    return None


def build_history_json(src_dirs, out_path: str) -> dict:
    """Parse every recognised .xlsx under one or more src dirs and write a
    consolidated, de-duplicated history.json. Returns a summary count dict."""
    if isinstance(src_dirs, str):
        src_dirs = [src_dirs]
    shifts: dict[tuple, dict] = {}
    deliveries: dict[tuple, dict] = {}
    fuel: dict[tuple, dict] = {}
    bales: dict[tuple, dict] = {}
    monthly: dict[str, dict] = {}
    errors: list[str] = []

    files = []
    for sd in src_dirs:
        files += glob.glob(os.path.join(sd, "**", "*.xlsx"), recursive=True)
    files = sorted(set(files))
    for path in files:
        if not os.path.isfile(path) or os.path.basename(path).startswith("~$"):
            continue  # skip directories named *.xlsx and Excel lock files
        kind = _classify(os.path.basename(path))
        if not kind:
            continue
        try:
            if kind == "tracker":
                for row in parse_tracker(path):
                    if row["date"][:7] < MIN_PERIOD:
                        continue
                    key = (row["date"], row["shift"])
                    # Keep the richer of any duplicate (handles "UPDATED" reissues
                    # and the cumulative tracker overlapping monthly ones).
                    prev = shifts.get(key)
                    if prev is None or row["qty"] > prev["qty"]:
                        shifts[key] = row
            elif kind == "deliveries":
                for row in parse_deliveries(path):
                    if row["date"][:7] < MIN_PERIOD:
                        continue
                    key = (row["date"], row["company"].lower(),
                           row["t30"], row["t12n"], row["t12ff"], row["pallets"])
                    deliveries[key] = row
            elif kind == "fuel":
                for row in parse_fuel(path):
                    if row["date"][:7] < MIN_PERIOD:
                        continue
                    key = (row["date"], row["shift"])
                    prev = fuel.get(key)
                    if prev is None or (row["usage"] or row["open"]) > (prev["usage"] or prev["open"]):
                        fuel[key] = row
            elif kind == "bales":
                for row in parse_bales(path):
                    if row["date"][:7] < MIN_PERIOD:
                        continue
                    key = (row["date"], row["grn"], row["quantity"])
                    bales[key] = row
            elif kind == "month_end":
                period = _month_end_period(path)
                if not period or period < MIN_PERIOD:
                    continue
                digest = parse_month_end(path, period)
                # Prefer the entry with the most populated detail.
                prev = monthly.get(period)
                if prev is None or len(json.dumps(digest)) > len(json.dumps(prev)):
                    monthly[period] = digest
        except Exception as e:  # keep going; report at the end
            errors.append(f"{os.path.basename(path)}: {e}")
            logger.warning("parse failed for %s: %s", path, e)

    data = {
        "min_period": MIN_PERIOD,
        "shifts": sorted(shifts.values(), key=lambda r: (r["date"], r["shift"])),
        "deliveries": sorted(deliveries.values(), key=lambda r: r["date"]),
        "fuel_dips": sorted(fuel.values(), key=lambda r: (r["date"], r["shift"])),
        "bales": sorted(bales.values(), key=lambda r: r["date"]),
        "monthly_stock": [monthly[p] for p in sorted(monthly)],
    }
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=1)
    summary = {k: len(v) for k, v in data.items() if isinstance(v, list)}
    summary["errors"] = errors
    return summary


# ---------------------------------------------------------------------------
# Load history.json into the database (idempotent)
# ---------------------------------------------------------------------------
def import_history(db, json_path: str | None = None) -> dict:
    """Load data/history.json into the DB, skipping anything already present.

    Idempotent: production shifts dedupe on (work_date, shift), deliveries on the
    full tuple, fuel on (date, shift), bales on (date, grn), monthly stock on
    period. Safe to run on every boot.
    """
    from datetime import datetime as _dt
    from ..models.production import ProductionShift, Shift
    from ..models.operations import Delivery, FuelDip, BaleReceipt, MonthlyStock
    from ..models.user import User, Role

    if json_path is None:
        here = os.path.dirname(__file__)
        json_path = os.path.normpath(os.path.join(here, "..", "..", "data", "history.json"))
    if not os.path.exists(json_path):
        logger.info("[history] no history.json at %s, skipping", json_path)
        return {}

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    admin = db.query(User).filter(User.role == Role.admin).first()
    admin_id = admin.id if admin else None
    shift_enum = {"Day": Shift.day, "Afternoon": Shift.afternoon, "Night": Shift.night}

    def d(s):
        return _dt.strptime(s, "%Y-%m-%d").date()

    counts = {"shifts": 0, "deliveries": 0, "fuel_dips": 0, "bales": 0, "monthly_stock": 0}

    existing = {(r.work_date, r.shift) for r in db.query(ProductionShift).all()}
    for r in data.get("shifts", []):
        wd, sh = d(r["date"]), shift_enum.get(r["shift"], Shift.night)
        if (wd, sh) in existing:
            continue
        existing.add((wd, sh))
        db.add(ProductionShift(
            work_date=wd, shift=sh, qty=r["qty"],
            p30s=r["p30s"], p30l=r["p30l"], p20n=r["p20n"], p12n=r["p12n"],
            p12hf=r["p12hf"], p12ff=r["p12ff"], p4cup=r["p4cup"], p2cup=r["p2cup"],
            hp1=r["hp1"], hp2=r["hp2"], hp3=r["hp3"], hp4=r["hp4"], hp5=r["hp5"], hp6=r["hp6"],
            labelling=r["label"], water_meter=r["water"], carton_bales=r["bales"],
            speed=r["speed"], fuel_open=r["fuel_open"], fuel_close=r["fuel_close"],
            fuel_use=r["fuel_use"], prod_hours=r["prod_hrs"], downtime_min=r["down_min"],
            sched_hours=r["tot_hrs"] or 8, clean_min=r["clean"], mold_min=r["mold"],
            other_min=r["other"], repulped=r["repulp"], comment=r["comment"],
            created_by=admin_id,
        ))
        counts["shifts"] += 1

    ex_del = {(x.work_date, (x.company or "").lower(), x.tray30, x.tray12n, x.tray12ff, x.pallets)
              for x in db.query(Delivery).all()}
    for r in data.get("deliveries", []):
        key = (d(r["date"]), r["company"].lower(), r["t30"], r["t12n"], r["t12ff"], r["pallets"])
        if key in ex_del:
            continue
        ex_del.add(key)
        db.add(Delivery(work_date=d(r["date"]), company=r["company"], tray30=r["t30"],
                        tray12n=r["t12n"], tray12ff=r["t12ff"], pallets=r["pallets"],
                        comment=r.get("comment", ""), created_by=admin_id))
        counts["deliveries"] += 1

    ex_fuel = {(x.work_date, x.shift) for x in db.query(FuelDip).all()}
    for r in data.get("fuel_dips", []):
        key = (d(r["date"]), r["shift"])
        if key in ex_fuel:
            continue
        ex_fuel.add(key)
        db.add(FuelDip(work_date=d(r["date"]), shift=r["shift"], open_dip=r["open"],
                       close_dip=r["close"], actual_usage=r["usage"], received=r["received"],
                       note=r["note"], created_by=admin_id))
        counts["fuel_dips"] += 1

    ex_bale = {(x.work_date, x.grn) for x in db.query(BaleReceipt).all()}
    for r in data.get("bales", []):
        key = (d(r["date"]), r["grn"])
        if key in ex_bale:
            continue
        ex_bale.add(key)
        db.add(BaleReceipt(work_date=d(r["date"]), grn=r["grn"], weight_kg=r["weight_kg"],
                           quantity=r["quantity"], created_by=admin_id))
        counts["bales"] += 1

    ex_stk = {x.period for x in db.query(MonthlyStock).all()}
    for r in data.get("monthly_stock", []):
        if r["period"] in ex_stk:
            continue
        ex_stk.add(r["period"])
        db.add(MonthlyStock(
            period=r["period"], diesel_eom=r["diesel_eom"], bal_30s=r["bal_30s"],
            bal_12n=r["bal_12n"], bal_12ff=r["bal_12ff"], bal_12nl=r["bal_12nl"],
            pallets_wrapped=r["pallets_wrapped"], bales_used=r["bales_used"],
            bales_purchased=r["bales_purchased"], labels_used=r["labels_used"],
            detail=r.get("detail"),
        ))
        counts["monthly_stock"] += 1

    db.commit()
    logger.info("[history] imported %s", counts)
    return counts


def main():
    import sys
    if len(sys.argv) >= 3 and sys.argv[1] == "build":
        summary = build_history_json(sys.argv[2], sys.argv[3])
        print("history.json built:", summary)
    else:
        from ..core.database import SessionLocal
        db = SessionLocal()
        try:
            print("imported:", import_history(db))
        finally:
            db.close()


if __name__ == "__main__":
    main()
