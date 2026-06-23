"""The AI improvement plan is embedded in the Excel and PDF reports when AI is
enabled, and absent (reports unchanged) when it's off. AI is mocked — no real
API call."""
import io

from openpyxl import load_workbook

FAKE = {
    "overall_summary": "Output steady; downtime the main drag.",
    "months": {},
    "improvement_plan": [
        {"title": "MOLD-WASH-DOWNTIME", "detail": "Top cause of lost time.",
         "action": "ACTION-TRIAL-MESH"},
        {"title": "Speed decline", "detail": "Trending down.", "action": "Check molds."},
    ],
}


def _seed(client, headers):
    for d in ("2023-05-08", "2023-05-18"):
        r = client.post("/api/v1/shifts", headers=headers,
                        json={"work_date": d, "shift": "Day", "qty": 900,
                              "fuel_use": 180, "downtime_min": 40, "sched_hours": 8,
                              "comment": "mold washing"})
        assert r.status_code in (201, 409), r.text


def _xlsx_sheets(content: bytes):
    wb = load_workbook(io.BytesIO(content))
    return wb, wb.sheetnames


def test_xlsx_omits_plan_when_ai_off(client, admin_headers):
    _seed(client, admin_headers)
    r = client.get("/api/v1/reports/report.xlsx", headers=admin_headers,
                   params={"start": "2023-05-01", "end": "2023-05-31"})
    assert r.status_code == 200, r.text
    _, sheets = _xlsx_sheets(r.content)
    assert "Improvement Plan" not in sheets
    assert "Summary" in sheets


def test_xlsx_includes_plan_when_ai_on(client, admin_headers, monkeypatch):
    _seed(client, admin_headers)
    monkeypatch.setattr("app.services.report_trends.ai_available", lambda: True)
    monkeypatch.setattr("app.services.report_trends.generate_deck_narrative", lambda payload: FAKE)
    r = client.get("/api/v1/reports/report.xlsx", headers=admin_headers,
                   params={"start": "2023-05-01", "end": "2023-05-31"})
    assert r.status_code == 200, r.text
    wb, sheets = _xlsx_sheets(r.content)
    assert "Improvement Plan" in sheets
    text = " ".join(str(c.value) for row in wb["Improvement Plan"].iter_rows()
                     for c in row if c.value is not None)
    assert "MOLD-WASH-DOWNTIME" in text
    assert "ACTION-TRIAL-MESH" in text


def test_pdf_builds_with_plan_when_ai_on(client, admin_headers, monkeypatch):
    _seed(client, admin_headers)
    monkeypatch.setattr("app.services.report_trends.ai_available", lambda: True)
    monkeypatch.setattr("app.services.report_trends.generate_deck_narrative", lambda payload: FAKE)
    r = client.get("/api/v1/reports/report.pdf", headers=admin_headers,
                   params={"start": "2023-05-01", "end": "2023-05-31"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF"
